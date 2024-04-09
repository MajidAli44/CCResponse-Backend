from tempfile import NamedTemporaryFile
from uuid import uuid4

import boto3
import openpyxl
import pdfkit
from boto3.s3.transfer import S3Transfer
from django.conf import settings
from django.template.loader import render_to_string


class CaseExportService:
    _empty_symbol = '-/-'

    @classmethod
    def upload_and_get_url(cls, file_path, time=600, key=None):
        """ upload a file to s3 and get signed url

        :param file_path: path of the file to upload
        :param time: expiration time of the generated url
        :param key: name of the file when stored in s3
        """
        ext = file_path.split('.')[-1]
        key = key or (str(uuid4()) + '.' + ext)
        client = boto3.client(
            's3', settings.AWS_S3_REGION_NAME,
            endpoint_url='https://ams3.digitaloceanspaces.com',
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY
        )
        transfer = S3Transfer(client)
        transfer.upload_file(file_path, settings.AWS_STORAGE_BUCKET_NAME, key)

        url = client.generate_presigned_url(
            'get_object',
            Params={'Bucket': settings.AWS_STORAGE_BUCKET_NAME, 'Key': key},
            ExpiresIn=time
        )
        return url

    @classmethod
    def export_to_excel(cls, queryset):
        work_b = openpyxl.Workbook()
        work_b.create_sheet(title='Case data', index=0)
        sheet = work_b['Case data']
        sheet.cell(row=1, column=1, value='Instruction date')
        sheet.cell(row=1, column=2, value='Accident date')
        sheet.cell(row=1, column=3, value='Client name')
        sheet.cell(row=1, column=4, value='CC Ref')
        sheet.cell(row=1, column=5, value='Phone number')
        sheet.cell(row=1, column=6, value='Introducer')
        sheet.cell(row=1, column=7, value='Provider')
        sheet.cell(row=1, column=8, value='Tp Insurer')
        sheet.cell(row=1, column=9, value='Instructed solicitor')
        for i, case in enumerate(queryset, start=2):
            sheet.cell(row=i, column=1, value=case.instruction_date.strftime('%d/%m/%Y'))
            if case.accident and case.accident.accident_date:
                sheet.cell(row=i, column=2, value=case.accident.accident_date.strftime('%d/%m/%Y'))
            else:
                sheet.cell(row=i, column=2, value=cls._empty_symbol)
            sheet.cell(row=i, column=3, value=case.client.name if case.client else cls._empty_symbol)
            sheet.cell(row=i, column=4, value=case.cc_ref)
            sheet.cell(row=i, column=5, value=case.client.phone_number if case.client else cls._empty_symbol)
            sheet.cell(row=i, column=6, value=case.introducer.name if case.introducer else cls._empty_symbol)
            sheet.cell(row=i, column=7, value=case.provider.name if case.provider else cls._empty_symbol)

            if case.third_party and case.third_party.insurer:
                sheet.cell(row=i, column=8, value=case.third_party.insurer.name if case.third_party.insurer else cls._empty_symbol)
            sheet.cell(row=i, column=9, value=case.solicitor.name if case.solicitor else cls._empty_symbol)

        with NamedTemporaryFile(delete=True, suffix='.xlsx') as tmp_file:
            work_b.save(tmp_file.name)
            url = cls.upload_and_get_url(tmp_file.name)

        return url

    @classmethod
    def export_to_pdf(cls, queryset):
        with NamedTemporaryFile(delete=True, suffix='.pdf') as tmp_file:
            pdfkit.from_string(render_to_string('export_pdf.html', context={'export_cases': queryset}), tmp_file.name)
            url = cls.upload_and_get_url(tmp_file.name)
            return url
